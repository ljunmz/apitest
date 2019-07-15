from openpyxl import load_workbook


def getTestCase(caseId, sheet):

    # # 接口excel用例，相对路径，也可以是绝对路径
    wb = load_workbook("F:\\code\\apitest\\resources\\apitestcase.xlsx")

    # 读取excel的table页，注意：你的excel，sheet页名称是什么这里就读什么
    sh = wb[sheet]

    # 用例总行数
    rows = sh.max_row
    caseList = []

    for item in range(2, rows + 1):
        caseIds = sh.cell(row=item, column=1).value
        if caseIds == caseId:
            '''excel第一列为字段名，所以我们要从第二行开始读    
            我们的列却是不变的，url、data、method    '''
            title = sh.cell(row=item, column=2).value
            url = sh.cell(row=item, column=3).value
            header = sh.cell(row=item, column=4).value
            data = sh.cell(row=item, column=5).value
            json = sh.cell(row=item, column=6).value
            method = sh.cell(row=item, column=7).value
            code = sh.cell(row=item, column=8).value
            msg = sh.cell(row=item, column=9).value
            # print(title, url, header, data, json, method, code, msg)
            caseJson = {"title": title, "url": url, "header": header, "data": data, "json": json, "method": method,
                        "code": code,
                        "msg": msg}
            caseList.append(caseJson)
    return caseList
